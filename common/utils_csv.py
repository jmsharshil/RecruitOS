import csv
import io
import re
from datetime import date, datetime
from decimal import Decimal
from django.http import HttpResponse
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font


def generate_csv_response(filename, headers, rows, export_format='csv', header_color=None, text_color=None):
    """
    Unified generator for CSV or XLSX export (called by *ExportView.get()).
    Defaults to CSV for backward compat. Uses openpyxl for XLSX with native Python
    types (bool, date, None) for proper cell typing in Excel. Updated buffer handling
    with seek(0) to prevent any truncation/corruption issues that could cause
    "file format or extension is not valid" errors. Supports ?format=xlsx and ?template=1.
    Cleans data per-row. See docs/common.md for details.
    """
    if export_format == 'xlsx':
        wb = Workbook()
        ws = wb.active
        ws.title = "Export"
        ws.append(headers)
        
        # Apply styles if colors are provided
        if header_color or text_color:
            fill = None
            font = None
            if header_color:
                h_col = str(header_color).lstrip('#')
                if len(h_col) == 6: h_col = 'FF' + h_col
                fill = PatternFill(start_color=h_col, end_color=h_col, fill_type='solid')
            if text_color:
                t_col = str(text_color).lstrip('#')
                if len(t_col) == 6: t_col = 'FF' + t_col
                font = Font(color=t_col, bold=True)
            else:
                font = Font(bold=True)
                
            for cell in ws[1]:
                if fill:
                    cell.fill = fill
                if font:
                    cell.font = font
                    
        for row in rows:
            # Clean row for Excel: use native types so openpyxl sets correct cell types
            # (boolean for bool, date serial for date, empty for None). Prevents corruption.
            clean_row = []
            for v in row:
                if v is None:
                    clean_row.append(None)
                elif isinstance(v, bool):
                    clean_row.append(v)
                elif isinstance(v, (date, datetime)):
                    clean_row.append(v)
                elif isinstance(v, Decimal):
                    clean_row.append(float(v))
                else:
                    clean_row.append(v)
            ws.append(clean_row)
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = filename.replace('.csv', '.xlsx').replace('.CSV', '.xlsx')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response.write(buffer.getvalue())
        return response
    else:
        # CSV fallback (original behavior)
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        writer = csv.writer(response)
        writer.writerow(headers)
        for row in rows:
            # Ensure bools/None are stringified nicely for CSV
            clean_row = ['' if v is None else str(v) for v in row]
            writer.writerow(clean_row)
        return response


def normalize_header(header):
    """Normalize header to snake_case for consistent key lookup (e.g. 'Candidate Name' -> 'candidate_name', 'Current CTC' -> 'current_ctc')."""
    if not header:
        return ''
    # Lower, replace non-alphanum with _, collapse multiples, strip
    normalized = re.sub(r'[^a-z0-9_]+', '_', str(header).lower().strip())
    normalized = re.sub(r'_+', '_', normalized).strip('_')
    return normalized


def parse_csv_from_request(request, required_fields=None):
    """
    Reads an uploaded CSV or Excel (.xlsx, .xls) file from request.FILES['file']
    and returns (normalized_headers, rows_as_dicts_with_norm_keys, error_string).
    Supports Excel via openpyxl.load_workbook(data_only=True) for computed values.
    - Header normalization (to snake_case) for both formats + required_fields validation.
    - CSV: utf-8-sig, DictReader.
    - Excel: iter_rows(min_row=2), skip fully empty rows, dict rows with stripped values.
    - Returns same contract as before for minimal caller impact.
    Updated error messages.
    """
    if 'file' not in request.FILES:
        return None, None, "No file uploaded. Send the file under the key 'file'."

    uploaded = request.FILES['file']
    filename = uploaded.name.lower()
    if not filename.endswith(('.csv', '.xlsx', '.xls')):
        return None, None, "Invalid file type. Only .csv, .xlsx, .xls files are accepted."

    try:
        if filename.endswith('.csv'):
            text = uploaded.read().decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(text))
            raw_headers = reader.fieldnames or []
            # Normalize headers for consistency (required_fields use snake_case)
            headers = [normalize_header(h) for h in raw_headers if h and str(h).strip()]
            rows = []
            for row in reader:
                if not any(str(v).strip() for v in row.values()):  # skip empty
                    continue
                norm_row = {}
                for raw_h, norm_h in zip(raw_headers, headers):
                    if norm_h:
                        val = row.get(raw_h, '') or ''
                        norm_row[norm_h] = str(val).strip()
                rows.append(norm_row)
        else:
            # Excel support
            uploaded.seek(0)
            workbook = openpyxl.load_workbook(uploaded, data_only=True)
            sheet = workbook.active
            raw_headers = []
            first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), [])
            for cell in first_row:
                if cell is not None:
                    cleaned = str(cell).strip()
                    if cleaned:
                        raw_headers.append(cleaned)
            headers = [normalize_header(h) for h in raw_headers]
            rows = []
            for row_tuple in sheet.iter_rows(min_row=2, values_only=True):
                if any(cell is not None and str(cell).strip() for cell in row_tuple if cell is not None):
                    row_dict = {}
                    for h, value in zip(headers, row_tuple):
                        if h:  # only add if normalized header
                            row_dict[h] = str(value).strip() if value is not None else ''
                    rows.append(row_dict)
    except Exception as e:
        file_type = 'CSV' if filename.endswith('.csv') else 'Excel'
        return None, None, f"Failed to parse {file_type} file: {str(e)}"

    if required_fields:
        missing = [f for f in required_fields if f not in headers]
        if missing:
            return None, None, f"Missing required columns: {', '.join(missing)}. Headers found: {headers}"

    return headers, rows, None


def get_choice(val, choices, default):
    """
    Case-insensitive helper to match a value against Django choices (value or display label).
    Used in CSV/Excel imports for status, priority, job_type etc. Falls back to default.
    See docs/candidates.md and docs/jobs.md for usage notes.
    """
    if not val:
        return default
    v = str(val).strip().lower()
    for c_val, c_disp in choices:
        if v in (str(c_val).lower(), str(c_disp).lower()):
            return c_val
    return default
